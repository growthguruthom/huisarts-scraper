"""Export results to Excel/CSV."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scraper.db import get_dashboard_data, get_unmatched_signalen


def export_excel(filename: str):
    """Export matches and unmatched signalen to an Excel file."""
    output_path = Path(__file__).parent.parent / "data" / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Gematchte resultaten
    ws_matches = wb.active
    ws_matches.title = "Matches"
    _write_matches_sheet(ws_matches)

    # Sheet 2: Ongematchte signalen
    ws_unmatched = wb.create_sheet("Ongematchte signalen")
    _write_unmatched_sheet(ws_unmatched)

    wb.save(str(output_path))

    # Also export CSV
    csv_path = output_path.with_suffix(".csv")
    _export_csv(csv_path)

    print(f"Geëxporteerd naar: {output_path}")
    print(f"CSV geëxporteerd naar: {csv_path}")


def _write_matches_sheet(ws):
    """Write matched results to a worksheet."""
    headers = [
        "Praktijknaam", "AGB-code", "Adres", "Postcode", "Stad",
        "Telefoon", "Website", "Signaaltype", "Signaal titel",
        "Omschrijving", "Bron URL", "Publicatiedatum", "Gemeente",
        "Match score", "Match type",
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    data = get_dashboard_data()
    for row_idx, row in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row.get("praktijk_naam"))
        ws.cell(row=row_idx, column=2, value=row.get("agb_code"))
        ws.cell(row=row_idx, column=3, value=row.get("praktijk_adres"))
        ws.cell(row=row_idx, column=4, value=row.get("praktijk_postcode"))
        ws.cell(row=row_idx, column=5, value=row.get("praktijk_stad"))
        ws.cell(row=row_idx, column=6, value=row.get("telefoon"))
        ws.cell(row=row_idx, column=7, value=row.get("website"))
        ws.cell(row=row_idx, column=8, value=row.get("signaal_type"))
        ws.cell(row=row_idx, column=9, value=row.get("signaal_titel"))
        ws.cell(row=row_idx, column=10, value=row.get("omschrijving"))
        ws.cell(row=row_idx, column=11, value=row.get("bron_url"))
        ws.cell(row=row_idx, column=12, value=row.get("publicatiedatum"))
        ws.cell(row=row_idx, column=13, value=row.get("gemeente"))
        ws.cell(row=row_idx, column=14, value=row.get("match_score"))
        ws.cell(row=row_idx, column=15, value=row.get("match_type"))

    # Auto-width columns
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    print(f"  Matches sheet: {len(data)} rijen")


def _write_unmatched_sheet(ws):
    """Write unmatched signalen to a worksheet."""
    headers = [
        "Type", "Titel", "Omschrijving", "Adres", "Postcode",
        "Stad", "Gemeente", "Bron URL", "Publicatiedatum",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    data = get_unmatched_signalen()
    for row_idx, row in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row.get("type"))
        ws.cell(row=row_idx, column=2, value=row.get("titel"))
        ws.cell(row=row_idx, column=3, value=row.get("omschrijving"))
        ws.cell(row=row_idx, column=4, value=row.get("adres"))
        ws.cell(row=row_idx, column=5, value=row.get("postcode"))
        ws.cell(row=row_idx, column=6, value=row.get("stad"))
        ws.cell(row=row_idx, column=7, value=row.get("gemeente"))
        ws.cell(row=row_idx, column=8, value=row.get("bron_url"))
        ws.cell(row=row_idx, column=9, value=row.get("publicatiedatum"))

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"
    print(f"  Ongematchte signalen sheet: {len(data)} rijen")


def _export_csv(path: Path):
    """Export matches as CSV."""
    import csv

    data = get_dashboard_data()
    if not data:
        return

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
